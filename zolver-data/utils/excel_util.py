import openpyxl
import pandas as pd
from pathlib import Path

class ExcelUtil:
    @staticmethod
    def load_properties(path: Path):
        wb = None
        try:
            # read_only=True / 메모리 효율성 확보
            wb = openpyxl.load_workbook(path, read_only=True)
            properties = wb.properties
            return properties
        except Exception as e:
            raise e
        finally:
            if 'wb' in locals() and wb:
                wb.close()
    
    @staticmethod
    def check_word_in_excel(file_path: Path, word_list: list) -> (bool):
        wb = None
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            found_words = set()
            total_to_find = len(word_list)

            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str):
                        for word in word_list:
                            if word in cell:
                                found_words.add(word)
                if len(found_words) == total_to_find:
                    return True
            return False
        except Exception as e:
            raise e
        finally:                
            if 'wb' in locals() and wb:
                wb.close()

    @staticmethod
    def read_excel_to_df(file_path: Path) -> pd.DataFrame:
        try:
            df = pd.read_excel(file_path, header=None)
            if df.empty and df is None:
                raise ValueError("EXCEL_IS_EMPTY")
                return pd.DataFrame()      
            return df  
        except Exception as e:
            raise e